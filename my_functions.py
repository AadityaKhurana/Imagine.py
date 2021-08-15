import requests
import pandas
import openpyxl


def get_fact():
    response = requests.get("https://uselessfacts.jsph.pl/today.json?language=en")
    json_response = response.json()
    fact = json_response["text"]
    return fact

def get_qoute():
    response = requests.get("https://quotes.rest/qod?language=en")
    json_response = response.json()
    quote = json_response["contents"]["quotes"][0]["quote"]
    author = json_response["contents"]["quotes"][0]["author"]
    return_list = [quote, author]
    return return_list

def reminder_check(name):
    my_dict = dict()
    db = pandas.read_excel('radio_btn.xlsx')
    my_array = db.to_numpy()
    for i in range(len(my_array)):
        my_dict[my_array[i][0]] = my_array[i][1]
    answer = my_dict[name]
    return answer

def set_reminder(name, input_value):
    my_dict = {'water': 2, "exercise": 3, "class": 4}
    path = "radio_btn.xlsx"
    wb_obj = openpyxl.load_workbook(path.strip())
    sheet_obj = wb_obj.active
    row = my_dict[name]
    obj_cell = sheet_obj.cell(row=row, column=2)
    obj_cell.value = input_value
    wb_obj.save("radio_btn.xlsx")

def check_time_table(name):
    my_dict_1 = dict()
    db = pandas.read_excel('time_table.xlsx')
    my_array = db.to_numpy()
    for i in range(len(my_array)):
        my_dict_1[my_array[i][0]] = my_array[i][1]
    return my_dict_1[name]

def set_time_table(slot, input_value):
    my_dict = {'9-10': 2, '10-11': 3, '11-12': 4, '12-13': 5, '13-14': 6, '14-15': 7, '15-16': 8, '16-17': 9, '17-18': 10, '18-19': 11, '19-20': 12, '20-21': 13}
    path = "time_table.xlsx"
    wb_obj = openpyxl.load_workbook(path.strip())
    sheet_obj = wb_obj.active
    row = my_dict[slot]
    obj_cell = sheet_obj.cell(row=row, column=2)
    obj_cell.value = input_value
    wb_obj.save("time_table.xlsx")

set_time_table('9-10', "new task")
print(check_time_table('9-10'))