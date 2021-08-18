import json
import requests
import pandas
import openpyxl
import random
import datetime

num = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]


def get_fact():
    response = requests.get(
        "https://uselessfacts.jsph.pl/today.json?language=en")
    json_response = response.json()
    fact = json_response["text"]
    return fact


def get_qoute():
    current_time = datetime.datetime.now()
    current_date = str(current_time.day) + '-' + \
        str(current_time.month) + '-' + str(current_time.year)
    db = pandas.read_excel('quote.xlsx')
    my_array = db.to_numpy()
    quote_added = bool
    for i in range(len(my_array)):
        if my_array[i][0] == current_date:
            print("Quote already there")
            quote = my_array[i][1]
            author = my_array[i][2]
            return_list = [quote, author]
            quote_added = True
        else:
            quote_added = False
    if quote_added == False:
        print("Fetching...")
        response = requests.get("https://quotes.rest/qod?language=en")
        json_response = response.json()
        quote = json_response["contents"]["quotes"][0]["quote"]
        author = json_response["contents"]["quotes"][0]["author"]
        return_list = [quote, author]
    return return_list


def get_youtube_vid():
    response = requests.get(
        "https://www.googleapis.com/youtube/v3/search?q=stand+up+comedy&key=AIzaSyD0VsHXytSG_QarxOHFzO9MDYMJydLT-Ag&maxResults=10")
    json_response = response.json()
    number = random.choice(num)
    vid_id = json_response["items"][number]['id']["videoId"]
    link = "https://www.youtube.com/embed/"+vid_id
    return link


def reminder_check(name):
    my_dict = dict()
    db = pandas.read_excel('radio_btn.xlsx')
    my_array = db.to_numpy()
    for i in range(len(my_array)):
        my_dict[my_array[i][0]] = my_array[i][1]
    answer = my_dict[name]
    return answer


def set_reminder(name, input_value):
    my_dict = {'water': 2, "meditate": 3,
               "exercise": 4, "rest": 5, "entertain": 6}
    path = "radio_btn.xlsx"
    wb_obj = openpyxl.load_workbook(path.strip())
    sheet_obj = wb_obj.active
    row = my_dict[name]
    obj_cell = sheet_obj.cell(row=row, column=2)
    obj_cell.value = input_value
    wb_obj.save("radio_btn.xlsx")


def check_time_table(name):
    my_dict_1 = dict()
    db = pandas.read_excel('time_table.xlsx')
    my_array = db.to_numpy()
    for i in range(len(my_array)):
        my_dict_1[my_array[i][0]] = my_array[i][1]
    return my_dict_1[name]


def set_time_table(slot, input_value):
    my_dict = {'9-10': 2, '10-11': 3, '11-12': 4, '12-13': 5, '13-14': 6, '14-15': 7,
               '15-16': 8, '16-17': 9, '17-18': 10, '18-19': 11, '19-20': 12, '20-21': 13}
    path = "time_table.xlsx"
    wb_obj = openpyxl.load_workbook(path.strip())
    sheet_obj = wb_obj.active
    row = my_dict[slot]
    obj_cell = sheet_obj.cell(row=row, column=2)
    obj_cell.value = input_value
    wb_obj.save("time_table.xlsx")
